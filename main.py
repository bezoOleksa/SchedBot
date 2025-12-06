import os
import sys
import time
import requests
import json
import openpyxl
import random
# import logging


ADMIN = os.getenv('TELEGRAM_BOT_ADMIN')
TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
API_URL = 'https://api.telegram.org/bot' + TOKEN

startMessage = 'Вітаю у боті для розкладу занять! 👋\n\n' \
               + 'Я буду надсилати вам автоматичні повідомлення на початку та в кінці уроку\n' \
               + 'Щоб вимкнути/увімкнути сповіщення, скористайтесь командами /mute та /unmute\n\n' \
               + 'Для отримання розкладу на сьогодні, надішліть /today\n' \
               + 'Для розкладу на завтра, надішліть /tomorrow\n\n' \
               + 'Для початку, давайте налаштуємо ваш розклад.'

helpMessage = '📚 Довідка по командах бота-розкладу\n\n' \
              + 'Це ваш помічник для швидкого доступу до розкладу занять та автоматичних сповіщень.\n\n' \
              + '--- Управління Розкладом ---\n' \
              + '/today або /sched — Показати ваш персональний розклад на поточний навчальний день.\n' \
              + '/tomorrow — Показати розклад на наступний навчальний день.\n\n' \
              + '--- Керування Сповіщеннями ---\n' \
              + '/unmute — Увімкнути автоматичні повідомлення про початок уроків та перерв.\n' \
              + '/mute — Вимкнути автоматичні повідомлення.\n\n' \
              + '--- Налаштування та Допомога ---\n' \
              + '/start — Розпочати або повторно пройти налаштування профілю (клас, група, підгрупа).\n' \
              + '/help — Показати це довідкове повідомлення.\n\n'

askRole = 'Будь ласка, оберіть вашу роль:'
keyboardRole = {'keyboard': [[{'text': 'Учень'}, {'text': 'Вчитель'}]],
                'resize_keyboard': True, 'one_time_keyboard': True}
answerRole = ('учень', 'вчитель')

askGrade = 'Чудово! Тепер оберіть ваш клас:'
keyboardGrade = {'keyboard': [[{'text': '9'}, {'text': '10'}, {'text': '11'}]],
                 'resize_keyboard': True, 'one_time_keyboard': True}

askGroup = 'Будь ласка, оберіть свою групу:'
groups = {'9': ['М-21', 'ІФ-22', 'ОІ-23', 'КМ-24', 'ПА-25'],
          '10': ['М-31', 'ІФ-32', 'ІЮ-33', 'КМ-34', 'ОІФ-35'],
          '11': ['М-41', 'ІФ-42', 'ПМ-43', 'ІН-ІФ-44']}

askHalf = 'І останнє, оберіть вашу підгрупу:'
keyboardHalf = {'keyboard': [[{'text': '1'}, {'text': '2'}]],
                'resize_keyboard': True, 'one_time_keyboard': True}

finalMessage = 'Дякую! Налаштування завершено. Тепер ви будете отримувати повідомлення з вашим персональним розкладом.'
teacherNote = 'Наразі функціонал для вчителів ще не розроблено, але ви все одно можете отримувати повідомлення про початок і кінець уроку. Дякуємо за розуміння!'
unrecognizedMessage = 'Вибачте, я не зрозумів вашу відповідь. Будь ласка, скористайтесь кнопками або командами. Для допомоги надішліть /help.'

scheduleSetupError = 'Будь ласка, спершу завершіть налаштування за допомогою команди /start, щоб отримати персональний розклад.'
weekendMessage = 'Цей день вихідний! Відпочивайте. 🥳'
schedForDay = '🗓️ Ваш розклад на '
todayUkr = 'сьогодні'
youAreHere = ' <-- 👈 Ви тут'
weekdaysUkr = ['понеділок', 'вівторок', 'середу', 'четвер', 'п\'ятницю', 'cуботу', 'неділю']

muteAnswer = '✅ Автоматичні сповіщення вимкнено. Щоб увімкнути їх знову, скористайтесь командою /unmute.'
unmuteAnswer = '✅ Автоматичні сповіщення увімкнено! Щоб вимкнути їх, скористайтесь командою /mute.'

lessonMessage = '🔔 Початок уроку'
breakMessage = '🎉 ПЕРЕРВА'
minsUkr = 'хв'
fiveMinsToStart = '🔔 Заняття розпочнуться через 5 хвилин! \n'
nextLessonUkr = 'Наступний урок: '
breakMotivMessages = [
    'Час для короткого відпочинку!',
    'Вдихни глибоко і розслабся!',
    'Насолодися моментом тиші.',
    'Відновлюй сили, попереду нові знання.',
    'Зроби перерву, ти на це заслуговуєш.',
    'Кілька хвилин для себе.',
    'Переключись на щось приємне.',
    'Подумай про щось приємне.',
    'Час для чаю або кави!',
    'Розслабся, скоро продовжимо.',
    'Невеличка пауза для великих звершень.'
]

endOfDayMessages = [
    'Це був останній урок на сьогодні! Вітаємо, ви впорались! 🎉',
    'Навчальний день завершено! Час відпочивати. ✨',
    'Уроки скінчились! Ви молодці! 👍',
    'Ще один день позаду! Гарного вечора!',
    'Ви чудово попрацювали! Тепер час для відпочинку.',
    'На сьогодні все! Набирайтесь сил на завтра.'
]

day = 60 * 60 * 24
twoMins = 2 * 60
timezonesDiff = 2 * 60 * 60
pollTimeout = 55


def loadFiles():
    global lastUpdate, TIMETABLE
    global Users
    try:
        with open('/storage/config.json', 'r', encoding='ascii') as file:
            data = json.load(file)
        lastUpdate = data['lastUpdate']
        TIMETABLE = data['TIMETABLE']
        print('Successfully loaded config.json')

        with open('/storage/users.json', 'r', encoding='utf-8') as file:
            Users = json.load(file)
        print('Successfully loaded users.json')

    except Exception as e:
        print('An error occurred while loading files (config.json, users.json):', e)
        sys.exit()


def makeSchedule():
    schedule = {}
    lessonsPerDay = 8
    daysPerWeek = 6
    startingRow = 13

    try:
        excelFile = openpyxl.load_workbook('/storage/schedule.xlsx')
        scheduleExcel = excelFile.active
        isMerged = lambda cell: isinstance(cell, openpyxl.cell.cell.MergedCell)
        groupNames = groups['9'] + groups['10'] + groups['11']  # grades 9, 10 and 11 possible only

        for group in range(2, 2 + len(groupNames) * 2):  # startingColumn = 2
            name = groupNames[group // 2 - 1]
            if name not in schedule:
                schedule[name] = [[], []]  # subgroups

            for day in range(daysPerWeek):
                schedule[name][group % 2].append([])

                for lesson in range(lessonsPerDay):
                    rowInExcel = startingRow + day * (lessonsPerDay + 1) + lesson
                    lessonCell = scheduleExcel.cell(row=rowInExcel, column=group)
                    if isMerged(lessonCell):
                        lessonCell = scheduleExcel.cell(row=rowInExcel, column=group - 1)
                    schedule[name][group % 2][day].append(lessonCell.value)

        with open('/storage/schedule.json', 'w', encoding='utf-8') as scheduleFile:
            json.dump(schedule, scheduleFile, indent=2, ensure_ascii=False)
        return schedule

    except Exception as e:
        print('Exception occured while making schedule:', e)
        sys.exit()


def sendMessage(chatID, text, keyboard={}):
    params = {'chat_id': chatID, 'text': text}
    if keyboard:
        params['reply_markup'] = json.dumps(keyboard)
    try:
        send = requests.post(API_URL + '/sendMessage', params=params, timeout=10)
        send.raise_for_status()
    except requests.exceptions.RequestException as e:
        print('Error ocurred sending message. Error:', e, end='; ')
        time.sleep(0.4 + random.random() / 2)
        try:
            send = requests.post(API_URL + '/sendMessage', params=params, timeout=10)
            send.raise_for_status()
            print('Sent on second attempt')
        except requests.exceptions.RequestException:
            print('No success on second attempt')


def uploadSchedule(document):
    global Schedule
    try:
        os.rename('/storage/schedule.xlsx', '/storage/schedule_backup.xlsx')
        fileID = document.get('file_id')
        getFile = requests.get(API_URL + '/getFile', params={'file_id': fileID})
        getFile.raise_for_status()
        filePath = getFile.json()['result']['file_path']

        newFile = requests.get(f'https://api.telegram.org/file/bot{TOKEN}/{filePath}')
        with open('/storage/schedule.xlsx', 'wb') as file:
            for chunk in newFile.iter_content(chunk_size=8192):
                if chunk:
                    file.write(chunk)

        Schedule = makeSchedule()
        sendMessage(ADMIN, 'Successfully updated schedule.xlsx file')
        os.remove('/storage/schedule_backup.xlsx')

    except Exception as e:
        os.remove('/storage/schedule.xlsx')
        os.rename('/storage/schedule_backup.xlsx', '/storage/schedule.xlsx')
        print('Error uploading schedule:', e)
        sendMessage(ADMIN, f'Error uploading schedule: {e}')


def uploadTimetable(document):
    global TIMETABLE, rerun
    try:
        prevTimetable = TIMETABLE
        fileID = document.get('file_id')
        getFile = requests.get(API_URL + '/getFile', params={'file_id': fileID})
        getFile.raise_for_status()
        filePath = getFile.json()['result']['file_path']
        newFile = requests.get(f'https://api.telegram.org/file/bot{TOKEN}/{filePath}')
        newFile.raise_for_status()
        TIMETABLE = json.loads(newFile.text)
        rerun = True

    except Exception as e:
        TIMETABLE = prevTimetable
        print('Error uploading timetable:', e)
        sendMessage(ADMIN, f'Error uploading timetable: {e}')


def makeTimePoints(now):  # Now
    global TimePoints
    TimePoints = []
    for timePoint in TIMETABLE:
        hour, min = timePoint.split(':')
        hour, min = int(hour), int(min)
        timeParams = (now.tm_year, now.tm_mon, now.tm_mday,
                      hour, min, 0, now.tm_wday, now.tm_yday, now.tm_isdst)
        TimePoints.append(time.struct_time(timeParams))


def notify():
    if Now.tm_wday in (5, 6):
        return

    for ID, user in Users.items():
        if not user.get('sendAuto'):
            continue

        if user.get('stage') < 5:
            if NextTimePoint == 0:
                sendMessage(ID, fiveMinsToStart)
            elif NextTimePoint % 2:
                sendMessage(ID, lessonMessage)
            else:
                sendMessage(ID, breakMessage)
            continue

        todaySched = Schedule[user['group']][user['half']][Now.tm_wday]
        if NextTimePoint == 0:
            sendMessage(ID, (fiveMinsToStart + '\n' if todaySched[0] else '') + makeDaySched(user))

        elif NextTimePoint % 2:
            if todaySched[NextTimePoint // 2]:
                sendMessage(ID, lessonMessage + ' ' + todaySched[NextTimePoint // 2])

        else:
            if (NextTimePoint == 14 or not todaySched[NextTimePoint // 2]) and todaySched[NextTimePoint // 2 - 1]:
                sendMessage(ID, random.choice(endOfDayMessages))

            elif todaySched[NextTimePoint // 2] and todaySched[NextTimePoint // 2 - 1]:
                breakStart, breakFinish = TimePoints[NextTimePoint], TimePoints[NextTimePoint + 1]
                breakDuration = breakFinish.tm_hour * 60 + breakFinish.tm_min - breakStart.tm_hour * 60 - breakStart.tm_min
                sendMessage(ID, f'{breakMessage} {breakDuration} {minsUkr}! {random.choice(breakMotivMessages)}\n'
                            + nextLessonUkr + todaySched[NextTimePoint // 2])

            else:
                continue


def makeDaySched(info, tomorrow=False):
    message = schedForDay + (weekdaysUkr[(Now.tm_wday + 1) % 7] if tomorrow else todayUkr) + ': \n'
    daySched = Schedule[info['group']][info['half']][(Now.tm_wday + tomorrow) % 7][:-1]

    for n, lesson in enumerate(daySched):
        message += f'\n{n + 1}. {TIMETABLE[2 * n + 1]}-{TIMETABLE[2 * n + 2]} - {lesson or "---"}'
        if not tomorrow and (n == (NextTimePoint - 1) // 2) and NextTimePoint:
            message += youAreHere
    return message


def reactToMessage(update):
    global UpdateUsers
    if 'message' not in update:
        return

    chatID = str(update['message']['chat']['id'])

    if 'document' in update['message'] and chatID == ADMIN:
        if update['message']['document'].get('file_name') == 'schedule.xlsx':
            uploadSchedule(update['message']['document'])
        elif update['message']['document'].get('file_name') == 'timetable.json':
            uploadTimetable(update['message']['document'])
        else:
            sendMessage(ADMIN, 'Please upload file named schedule.xlsx to update schedule; ' \
                             + 'timetable.json to update timetable')
            return

    if 'text' not in update['message']:
        return
    text = update['message']['text']

    if text == '/start' or chatID not in Users:
        Users[chatID] = {'sendAuto': True, 'stage': 0}
        sendMessage(chatID, startMessage)
        sendMessage(chatID, askRole, keyboard=keyboardRole)
        Users[chatID]['stage'] = 1
        UpdateUsers = True

    elif text == '/help':
        sendMessage(chatID, helpMessage)

    elif text == '/mute':
        Users[chatID]['sendAuto'] = False
        sendMessage(chatID, muteAnswer)
        UpdateUsers = True

    elif text == '/unmute':
        Users[chatID]['sendAuto'] = True
        sendMessage(chatID, unmuteAnswer)
        UpdateUsers = True

    elif text in ('/sched', '/today', '/tomorrow'):
        info = Users.get(chatID)
        if not info or info.get('stage', 0) < 5:
            sendMessage(chatID, scheduleSetupError)
            return

        tomorrow = text == '/tomorrow'
        if (Now.tm_wday + tomorrow) % 7 == 6:
            sendMessage(chatID, weekendMessage)
            return

        sendMessage(chatID, makeDaySched(info, tomorrow=tomorrow))

    else:
        stage = Users.get(chatID, {}).get('stage')
        UpdateUsers = True
        if stage == 1 and text.lower() in answerRole:
            if text.lower() == answerRole[0]:
                Users[chatID]['role'] = 'student'
                sendMessage(chatID, askGrade, keyboard=keyboardGrade)
                Users[chatID]['stage'] = 2
            else:
                Users[chatID]['role'] = 'teacher'
                sendMessage(chatID, teacherNote)
                sendMessage(chatID, askRole, keyboard=keyboardRole)
                Users[chatID]['stage'] = 1

        elif stage == 2 and text in groups:
            Users[chatID]['grade'] = text
            keyboardGroup = {'keyboard': [[{'text': group} for group in groups[text]]],
                             'resize_keyboard': True, 'one_time_keyboard': True}
            sendMessage(chatID, askGroup, keyboard=keyboardGroup)
            Users[chatID]['stage'] = 3

        elif stage == 3 and text in groups.get(Users[chatID].get('grade'), []):
            Users[chatID]['group'] = text
            sendMessage(chatID, askHalf, keyboard=keyboardHalf)
            Users[chatID]['stage'] = 4

        elif stage == 4 and text in ('1', '2'):
            Users[chatID]['half'] = int(text) - 1
            sendMessage(chatID, finalMessage, keyboard={'remove_keyboard': True})
            Users[chatID]['stage'] = 5

        else:
            UpdateUsers = False
            sendMessage(chatID, unrecognizedMessage)


def getUpdates(offset=None, timeout=pollTimeout):
    try:
        params = {'offset': offset, 'timeout': timeout}
        response = requests.get(API_URL + '/getUpdates', params=params, timeout=timeout + 10)
        response.raise_for_status()
        return response.json()['result']
    except requests.exceptions.RequestException as e:
        print('Error getting updates: ', e)
        return []


def saveToFiles():
    global UpdateUsers
    try:
        with open('/storage/config.json', 'w', encoding='ascii') as configFile:
            json.dump({'lastUpdate': lastUpdate, 'TIMETABLE': TIMETABLE}, configFile, indent=2)

        if UpdateUsers:
            with open('/storage/users.json', 'w', encoding='utf-8') as usersFile:
                json.dump(Users, usersFile, indent=2, ensure_ascii=False)
            UpdateUsers = False

    except Exception as e:
        print('Error occurred while trying to save config.json or/and users.json', e)


if __name__ == '__main__':
    rerun = True
    while rerun:
        rerun = False
        Now = time.localtime(time.time() + timezonesDiff)
        TimePoints = []
        NextTimePoint = 0
        lastUpdate = None
        TIMETABLE = []
        Users = {}
        UpdateUsers = False

        loadFiles()
        Schedule = makeSchedule()
        makeTimePoints(Now)

        current_time_mk = time.mktime(Now)
        while NextTimePoint < 15 and current_time_mk >= time.mktime(TimePoints[NextTimePoint]) + twoMins:
            NextTimePoint += 1

        if NextTimePoint >= 15:
            makeTimePoints(time.localtime(current_time_mk + day))
            NextTimePoint = 0

        print(f"{Now.tm_mday}.{Now.tm_mon}.{Now.tm_year} {Now.tm_hour}:{Now.tm_min}:{Now.tm_sec} Bot started")
        sendMessage(ADMIN, 'Bot started')

        while True:
            Now = time.localtime(time.time() + timezonesDiff)
            current_time_mk = time.mktime(Now)
            timeToNextEvent = time.mktime(TimePoints[NextTimePoint]) - current_time_mk

            if timeToNextEvent <= 0:
                if abs(timeToNextEvent) < twoMins:
                    notify()

                if NextTimePoint < 14:
                    NextTimePoint += 1
                else:
                    makeTimePoints(time.localtime(current_time_mk + day))
                    NextTimePoint = 0

            updates = getUpdates(lastUpdate, max(5, min(pollTimeout, int(timeToNextEvent))))
            for update in updates:
                reactToMessage(update)
                lastUpdate = update['update_id'] + 1

            saveToFiles()
            if rerun:
                break
